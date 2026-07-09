# =============================================================================
# TARV-Score — Application de scoring clinique | Bilingual FR/EN
# Mémoire ISE3 | ISSEA-CEMAC 2025-2026 | CNLS / GTC / Cameroun
# Auteur : AZONFACK MYRIAM DOLVIANNE
# Modèle : meilleur des 4 candidats (Rég. Logistique / Random Forest / SVM /
#          XGBoost), sélectionné sur score composite (6 métriques à poids égal),
#          entraîné en Class Weight sur les 14 variables les plus explicatives
#          de l'interruption au TARV (cf. train_model.py).
# =============================================================================

import base64
import json
import math
from datetime import datetime
from pathlib import Path

import joblib
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from fpdf import FPDF

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TARV-Score | CNLS Cameroun",
    page_icon="🎗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# TRADUCTIONS FR / EN
# ─────────────────────────────────────────────────────────────────────────────
T = {
    # ── Sidebar ──────────────────────────────────────────────────────────────
    "lang_label":    {"fr": "🌐 Langue / Language", "en": "🌐 Language / Langue"},
    "lang_fr":       {"fr": "🇫🇷 Français",          "en": "🇫🇷 French"},
    "lang_en":       {"fr": "🇬🇧 Anglais",           "en": "🇬🇧 English"},

    "sidebar_title": {"fr": "🎗️ TARV-Score",         "en": "🎗️ TARV-Score"},
    "sidebar_sub":   {"fr": "CNLS Cameroun · ISSEA-CEMAC 2025-2026",
                      "en": "CNLS Cameroon · ISSEA-CEMAC 2025-2026"},
    "perf_title":    {"fr": "ℹ️ À propos de l'outil",
                      "en": "ℹ️ About this tool"},
    "perf_body":     {
        "fr": "Basé sur le suivi de **2 720 patients** par le CNLS en 2024. Méthode de calcul validée statistiquement.",
        "en": "Based on the follow-up of **2,720 patients** by NACC in 2024. Statistically validated calculation method.",
    },
    "seuil_lbl":     {"fr": "Seuil",                 "en": "Threshold"},

    "grid_title":    {"fr": "🚦 Grille d'interprétation",
                      "en": "🚦 Risk Interpretation Grid"},
    "risk_low":      {"fr": "Faible",                "en": "Low"},
    "risk_mod":      {"fr": "Modéré",                "en": "Moderate"},
    "risk_high":     {"fr": "Élevé",                 "en": "High"},
    "risk_low_desc": {"fr": "probabilité < 30 %",    "en": "probability < 30%"},
    "risk_mod_desc": {"fr": "probabilité 30 % – 50 %", "en": "probability 30%–50%"},
    "risk_high_desc":{"fr": "probabilité ≥ 50 %",   "en": "probability ≥ 50%"},

    "ctx_title":     {"fr": "ℹ️ Contexte de l'étude",
                      "en": "ℹ️ Study Context"},
    "ctx_body":      {
        "fr": ("Données CNLS 2024 · Enquête nationale auprès de **2 720 PvVIH** dans les 10 régions du Cameroun.\n\n"
               "Outil basé sur **14 caractéristiques du patient** (parcours de soins, suivi thérapeutique, "
               "profil socio-démographique), choisies pour leur lien statistique avec l'interruption du TARV "
               "et validées avec le CNLS."),
        "en": ("CNLS 2024 data · National survey of **2,720 PLHIV** across Cameroon's 10 regions.\n\n"
               "Tool based on **14 patient characteristics** (care pathway, therapeutic follow-up, "
               "socio-demographic profile), chosen for their statistical link with ART interruption "
               "and validated with NACC."),
    },
    "disclaimer":    {
        "fr": "⚠️ Outil d'aide à la décision.<br>Ne remplace pas le jugement clinique.",
        "en": "⚠️ Decision-support tool.<br>Does not replace clinical judgment.",
    },

    # ── En-tête ──────────────────────────────────────────────────────────────
    "main_title":    {
        "fr": "OUTIL DE DÉTECTION DU RISQUE D'INTERRUPTION AU TARV<br>CHEZ LES PVVIH AU CAMEROUN",
        "en": "RISK OF ART INTERRUPTION DETECTION TOOL<br>AMONG PLHIV IN CAMEROON",
    },
    "main_sub":      {
        "fr": "Prédiction du risque d'interruption du Traitement Antirétroviral",
        "en": "Prediction of the risk of Antiretroviral Treatment interruption",
    },
    "badge_cnls":    {"fr": "CNLS Cameroun",         "en": "NACC Cameroon"},
    "badge_validated": {"fr": "✓ Outil validé",       "en": "✓ Validated tool"},

    # ── Formulaire ───────────────────────────────────────────────────────────
    "form_intro":    {
        "fr": "📋 Profil du patient — 14 caractéristiques à renseigner",
        "en": "📋 Patient Profile — 14 characteristics to fill in",
    },
    "form_sub":      {
        "fr": "Renseignez toutes les caractéristiques, puis cliquez sur <strong>Calculer le Score</strong>.",
        "en": "Fill in all characteristics, then click <strong>Calculate Score</strong>.",
    },
    "sec_soins":     {"fr": "📍 Parcours de soins",
                      "en": "📍 Care Pathway"},
    "sec_thera":     {"fr": "💊 Suivi thérapeutique",
                      "en": "💊 Therapeutic Follow-up"},
    "sec_socio":     {"fr": "👤 Profil socio-démographique",
                      "en": "👤 Socio-demographic Profile"},

    # Variables
    "region":        {"fr": "Région",                "en": "Region"},
    "type_fosa":     {"fr": "Type de FOSA",          "en": "Health Facility Type"},
    "dsd":           {"fr": "Mode de dispensation (DSD)",
                      "en": "Dispensation Mode (DSD)"},
    "pepfar":        {"fr": "Soutien PEPFAR",        "en": "PEPFAR Support"},
    "observance":    {"fr": "Observance (4 derniers jours)",
                      "en": "Adherence (last 4 days)"},
    "retesting":     {"fr": "Nouveau dépistage VIH (retesting)",
                      "en": "HIV Retesting"},
    "sexe":          {"fr": "Sexe",                  "en": "Sex"},
    "tranche_age":   {"fr": "Tranche d'âge",         "en": "Age Group"},
    "statut_mat":    {"fr": "Statut matrimonial",    "en": "Marital Status"},
    "religion":      {"fr": "Religion",              "en": "Religion"},
    "niveau_etude":  {"fr": "Niveau d'étude",        "en": "Education Level"},
    "depenses":      {"fr": "Dépenses mensuelles liées au traitement",
                      "en": "Monthly Treatment-related Expenses"},
    "soutien_familial": {"fr": "Soutien familial",   "en": "Family support"},
    "traitement_alt":   {"fr": "Recours à un traitement alternatif",
                          "en": "Use of alternative treatment"},

    "oui":           {"fr": "Oui",                   "en": "Yes"},
    "non":           {"fr": "Non",                   "en": "No"},
    "masculin":      {"fr": "Masculin",              "en": "Male"},
    "feminin":       {"fr": "Féminin",               "en": "Female"},

    "btn_calc":      {"fr": "🔍  Calculer le Score de Risque",
                      "en": "🔍  Calculate Risk Score"},
    "spinner":       {"fr": "Calcul du score en cours…",
                      "en": "Computing score…"},

    # ── Résultats ────────────────────────────────────────────────────────────
    "res_divider":   {"fr": "📊 Résultat du Scoring Clinique",
                      "en": "📊 Clinical Scoring Result"},
    "prob_label":    {"fr": "probabilité d'interruption",
                      "en": "interruption probability"},
    "seuil_txt":     {"fr": "Seuil de décision",     "en": "Decision threshold"},
    "reco_lbl":      {"fr": "Recommandation clinique :",
                      "en": "Clinical Recommendation:"},
    "risk_low_lbl":  {"fr": "FAIBLE",                "en": "LOW"},
    "risk_mod_lbl":  {"fr": "MODÉRÉ",                "en": "MODERATE"},
    "risk_high_lbl": {"fr": "ÉLEVÉ",                 "en": "HIGH"},

    "reco_low": {
        "fr": ("Profil stable. Maintenir le suivi standard. Renforcer les messages "
               "d'observance lors de la prochaine dispensation."),
        "en": ("Stable profile. Maintain standard follow-up. Reinforce adherence "
               "counselling at next dispensation visit."),
    },
    "reco_mod": {
        "fr": ("Suivi renforcé recommandé. Identifier les freins à l'observance "
               "(transport, soutien familial, effets indésirables) et proposer "
               "un accompagnement individualisé dès cette consultation."),
        "en": ("Enhanced follow-up recommended. Identify barriers to adherence "
               "(transport, family support, side effects) and offer individualised "
               "support starting from this visit."),
    },
    "reco_high": {
        "fr": ("Intervention prioritaire requise. Déclencher un plan de rétention : "
               "contact téléphonique dans les 48 h, visite à domicile (VAD), "
               "orientation vers un accompagnement psychosocial et révision "
               "du mode de dispensation (DSD) si possible."),
        "en": ("Priority intervention required. Launch a retention plan: "
               "phone contact within 48 h, home visit (VAD), referral for "
               "psychosocial support, and review of the dispensation mode (DSD) if possible."),
    },

    "imp_sub":       {
        "fr": "Plus la barre est longue, plus la caractéristique pèse dans le calcul du score.",
        "en": "Longer bar = more weight in the score calculation.",
    },
    "imp_title":     {"fr": "🔬 Top {} facteurs les plus contributifs",
                      "en": "🔬 Top {} most contributing factors"},
    "imp_xlabel":    {"fr": "Poids dans le calcul du score",
                      "en": "Weight in the score calculation"},
    "imp_legend":    {
        "fr": "🔴 Importance forte &nbsp;|&nbsp; 🟠 Modérée &nbsp;|&nbsp; 🟢 Modérée-faible &nbsp;|&nbsp; 🔵 Faible",
        "en": "🔴 High importance &nbsp;|&nbsp; 🟠 Moderate &nbsp;|&nbsp; 🟢 Moderate-low &nbsp;|&nbsp; 🔵 Low",
    },
    "recap_title":   {"fr": "🗂️ Récapitulatif du profil saisi",
                      "en": "🗂️ Summary of entered profile"},
    "recap_var":     {"fr": "Variable",              "en": "Variable"},
    "recap_val":     {"fr": "Valeur",                "en": "Value"},

    # ── Pied de page ─────────────────────────────────────────────────────────
    "footer": {
        "fr": ("🎗️ <strong>TARV-Score</strong> — Mémoire ISE3 ISSEA-CEMAC 2025-2026 &nbsp;|&nbsp; "
               "CNLS / GTC Cameroun<br>"
               "<em>⚠️ Outil d'aide à la décision — Ne remplace pas le jugement clinique du prestataire de santé.</em>"),
        "en": ("🎗️ <strong>TARV-Score</strong> — ISE3 Thesis ISSEA-CEMAC 2025-2026 &nbsp;|&nbsp; "
               "NACC / GTC Cameroon<br>"
               "<em>⚠️ Decision-support tool — Does not replace the clinical judgment of the healthcare provider.</em>"),
    },

    # ── Options variables ─────────────────────────────────────────────────────
    "region_opts": {
        "fr": ["Centre","Adamaoua","Est","Extrême-Nord","Littoral",
               "Nord","Nord-Ouest","Ouest","Sud","Sud-Ouest"],
        "en": ["Centre","Adamawa","East","Far North","Littoral",
               "North","North-West","West","South","South-West"],
    },
    "type_fosa_opts": {
        "fr": ["Public","Privé confessionnel","Privé laïc"],
        "en": ["Public","Faith-based Private","Secular Private"],
    },
    "dsd_opts": {
        "fr": ["Standard","DSD avec décalage RDV","DSD sans décalage"],
        "en": ["Standard","DSD with appointment delay","DSD without delay"],
    },
    "observance_opts": {
        "fr": ["Bonne","Modérée","Médiocre"],
        "en": ["Good","Moderate","Poor"],
    },
    "religion_opts": {
        "fr": ["Catholique","Protestant","Musulman","Autre"],
        "en": ["Catholic","Protestant","Muslim","Other"],
    },
    "niveau_etude_opts": {
        "fr": ["Primaire","Jamais fréquenté","Secondaire Premier Cycle",
               "Secondaire Second Cycle","Supérieur"],
        "en": ["Primary","Never attended school","Lower Secondary",
               "Upper Secondary","Higher Education"],
    },
    "depenses_opts": {
        "fr": ["Moins de 5 000","[5 000 - 10 000[","[10 000 - 25 000[","25 000 et plus"],
        "en": ["Under 5,000 FCFA","5,000-10,000 FCFA","10,000-25,000 FCFA","25,000 FCFA and above"],
    },
    "tranche_age_opts": {
        "fr": ["25 à 49 Ans","18 à 20 Ans","21 à 24 Ans","50 ans et plus"],
        "en": ["25 to 49 years","18 to 20 years","21 to 24 years","50 years and above"],
    },
    "statut_mat_opts": {
        "fr": ["Marié(e) en monogamie","Célibataire","En union libre/concubinage",
               "Marié(e) en polygamie","En séparation de corps / Divorcée","Veuf (ve)"],
        "en": ["Married (monogamous)","Single","Cohabiting/Common-law",
               "Married (polygamous)","Separated / Divorced","Widowed"],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# CSS GLOBAL
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #eef2f7; }

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0b2d52 0%, #1a5e8a 55%, #0d7a5c 100%);
}
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #ddeeff; }
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 { color: #ffffff !important; font-weight: 700 !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.2) !important; }
section[data-testid="stSidebar"] [data-testid="metric-container"] {
    background: rgba(255,255,255,0.12); border-radius: 8px; padding: 8px 12px;
}
section[data-testid="stSidebar"] [data-testid="metric-container"] label { color: #a8d0f0 !important; }
section[data-testid="stSidebar"] [data-testid="stMetricValue"]
    { color: #fff !important; font-size: 1.3em !important; font-weight: 700 !important; }

div[data-testid="stForm"] button[kind="primaryFormSubmit"],
button[kind="primary"] {
    background: linear-gradient(135deg, #0b2d52 0%, #1a6fa8 100%) !important;
    border: none !important; color: #fff !important;
    font-weight: 700 !important; font-size: 1.05em !important;
    border-radius: 10px !important; padding: 14px !important;
    box-shadow: 0 4px 16px rgba(11,45,82,0.35) !important;
    letter-spacing: 0.3px !important; transition: all 0.2s !important;
}
button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 7px 22px rgba(11,45,82,0.45) !important;
}
.stSelectbox label, .stRadio label {
    font-size: 0.87em !important; font-weight: 600 !important; color: #1a2d4a !important;
}
.res-divider {
    display: flex; align-items: center; gap: 14px; margin: 24px 0 18px 0;
}
.res-divider hr { flex: 1; border: none; border-top: 2px solid #ccd8e8; }
.res-divider span {
    background: linear-gradient(135deg, #0b2d52, #1a6fa8);
    color: #fff; border-radius: 25px; padding: 5px 22px;
    font-size: 0.86em; font-weight: 700; white-space: nowrap;
    box-shadow: 0 3px 10px rgba(11,45,82,0.25);
}
.footer {
    text-align: center; font-size: 0.75em; color: #8a9bb0;
    margin-top: 30px; padding: 14px 0;
    border-top: 1px solid #ccd8e8; line-height: 1.8;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CHEMINS & RESSOURCES
# ─────────────────────────────────────────────────────────────────────────────
BASE    = Path(__file__).parent
ASSETS  = BASE / "assets"
MODELES = BASE / "modeles"


@st.cache_resource
def load_resources():
    model      = joblib.load(MODELES / "modele_final.pkl")
    scaler     = joblib.load(MODELES / "scaler.pkl")
    clean_cols = joblib.load(MODELES / "colonnes_modele.pkl")
    references = joblib.load(MODELES / "references.pkl")
    variables  = joblib.load(MODELES / "variables_modele.pkl")
    with open(MODELES / "meta.json", encoding="utf-8") as f:
        meta = json.load(f)
    return model, scaler, clean_cols, references, variables, meta


def img_b64(path: Path) -> str:
    with open(path, "rb") as fh:
        return base64.b64encode(fh.read()).decode()


# ─────────────────────────────────────────────────────────────────────────────
# SÉLECTEUR DE LANGUE (sidebar, en premier)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"### {T['lang_label']['fr']}")
    lang_choice = st.radio(
        label="",
        options=["fr", "en"],
        format_func=lambda x: T["lang_fr"][x] if x == "fr" else T["lang_en"][x],
        horizontal=True,
        key="lang",
        label_visibility="collapsed",
    )

L = lang_choice  # "fr" ou "en"


def t(key: str) -> str:
    return T[key][L]


# ─────────────────────────────────────────────────────────────────────────────
# CHARGEMENT DES RESSOURCES DU MODELE
# ─────────────────────────────────────────────────────────────────────────────
try:
    model, scaler, clean_cols, references, VARS_FIN, meta = load_resources()
    SCALER_COLS = list(scaler.feature_names_in_)
except Exception as exc:
    st.error(f"❌ {'Impossible de charger le modèle' if L=='fr' else 'Cannot load the model'}: {exc}")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# SEUIL DE DECISION (sidebar)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"## {t('sidebar_title')}")
    st.caption(t("sidebar_sub"))
    st.divider()

    st.markdown(f"### {t('perf_title')}")
    st.markdown(f'<p style="font-size:0.85em;line-height:1.5;opacity:0.9;">{t("perf_body")}</p>',
                unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    seuil_title = "🎚️ Seuil de décision" if L == "fr" else "🎚️ Decision Threshold"
    st.markdown(f"**{seuil_title}**")
    SEUIL = st.slider(
        label="",
        min_value=0.20,
        max_value=0.80,
        value=float(meta["seuil_defaut"]),
        step=0.01,
        format="%.2f",
        key="seuil_slider",
        label_visibility="collapsed",
    )
    st.caption(f"{t('seuil_lbl')} : {SEUIL:.0%}")

    st.divider()
    st.markdown(f"### {t('grid_title')}")
    grid_intro = ("Le score est interprété selon 3 niveaux de risque :"
                  if L == "fr" else
                  "The score is interpreted across 3 risk levels:")
    st.markdown(f'<p style="font-size:0.82em;opacity:0.8;margin-bottom:8px;">{grid_intro}</p>',
                unsafe_allow_html=True)
    for emoji, key_lbl, key_desc, border_col, bg_col, txt_col in [
        ("🟢", "risk_low",  "risk_low_desc",  "#27ae60", "rgba(39,174,96,0.18)",  "#90ffb0"),
        ("🟡", "risk_mod",  "risk_mod_desc",  "#f1c40f", "rgba(241,196,15,0.18)", "#fff176"),
        ("🔴", "risk_high", "risk_high_desc", "#e74c3c", "rgba(231,76,60,0.18)",  "#ffaaaa"),
    ]:
        st.markdown(
            f'<div style="border-left:4px solid {border_col};background:{bg_col};'
            f'border-radius:0 8px 8px 0;padding:9px 12px;margin-bottom:7px;">'
            f'<div style="color:{txt_col};font-weight:700;font-size:0.9em;">{emoji} {t(key_lbl)}</div>'
            f'<div style="color:rgba(255,255,255,0.75);font-size:0.8em;margin-top:2px;">{t(key_desc)}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.divider()
    st.markdown(f"### {t('ctx_title')}")
    st.markdown(t("ctx_body"))
    st.divider()
    st.markdown(
        f'<div style="font-size:0.74em;opacity:0.65;text-align:center;line-height:1.7;">'
        f'{t("disclaimer")}</div>',
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# EN-TÊTE PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
cnls_b64  = img_b64(ASSETS / "cnls_logo.png")
issea_b64 = img_b64(ASSETS / "issea_logo.png")

st.markdown(f"""
<div style="
    background: linear-gradient(135deg, #0b2d52 0%, #1a5e8a 50%, #0d7a5c 100%);
    border-radius: 18px; padding: 22px 36px; margin-bottom: 26px;
    box-shadow: 0 10px 36px rgba(11,45,82,0.30);
    display: flex; align-items: center; justify-content: space-between; gap: 20px;">
    <img src="data:image/png;base64,{cnls_b64}" height="130"
         style="border-radius:12px; box-shadow:0 6px 20px rgba(0,0,0,0.40); flex-shrink:0; object-fit:contain;">
    <div style="text-align:center; color:#fff; flex:1;">
        <div style="font-size:3.8em; line-height:1; margin-bottom:8px;">
            <span style="color:#ff2222; filter:drop-shadow(0 0 10px #ff000099) drop-shadow(0 0 4px #ff0000cc);">🎗️</span>
        </div>
        <div style="font-size:1.22em; font-weight:900; letter-spacing:1px; text-shadow:0 2px 8px rgba(0,0,0,0.35); line-height:1.3; text-transform:uppercase;">{t("main_title")}</div>
        <div style="font-size:0.88em; font-weight:300; opacity:0.88; margin-top:9px;">{t("main_sub")}</div>
        <div style="display:flex; gap:8px; justify-content:center; flex-wrap:wrap; margin-top:13px;">
            <span style="background:rgba(255,255,255,0.15);border:1px solid rgba(255,255,255,0.3);border-radius:20px;padding:3px 14px;font-size:0.78em;font-weight:600;">{t("badge_cnls")}</span>
            <span style="background:rgba(39,174,96,0.35);border:1px solid rgba(39,174,96,0.6);border-radius:20px;padding:3px 14px;font-size:0.78em;font-weight:600;">{t("badge_validated")}</span>
        </div>
    </div>
    <img src="data:image/png;base64,{issea_b64}" height="130"
         style="border-radius:12px; box-shadow:0 6px 20px rgba(0,0,0,0.40); flex-shrink:0; object-fit:contain;">
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# MAPPINGS INTERNES (valeurs d'entraînement — NE PAS MODIFIER)
# ─────────────────────────────────────────────────────────────────────────────
FOSA_INTERNAL = {
    "Public": "Public",
    "Privé confessionnel": "Privé confessionnel",
    "Privé laïc": "Privé laic",            # pas de ï dans les données
    "Faith-based Private": "Privé confessionnel",
    "Secular Private": "Privé laic",
}
OBSERVANCE_INTERNAL = {
    "Bonne": "Bonne",  "Modérée": "Modérée",  "Médiocre": "Mediocre",  # pas de é
    "Good": "Bonne",   "Moderate": "Modérée",  "Poor": "Mediocre",
}

REGION_EN2FR   = dict(zip(T["region_opts"]["en"],       T["region_opts"]["fr"]))
ETUDE_EN2FR    = dict(zip(T["niveau_etude_opts"]["en"], T["niveau_etude_opts"]["fr"]))
DSD_EN2FR      = dict(zip(T["dsd_opts"]["en"],          T["dsd_opts"]["fr"]))
RELIGION_EN2FR = dict(zip(T["religion_opts"]["en"],     T["religion_opts"]["fr"]))
DEPENSES_EN2FR = dict(zip(T["depenses_opts"]["en"],     T["depenses_opts"]["fr"]))
TRANCHE_EN2FR  = dict(zip(T["tranche_age_opts"]["en"],  T["tranche_age_opts"]["fr"]))
STATUT_EN2FR   = dict(zip(T["statut_mat_opts"]["en"],   T["statut_mat_opts"]["fr"]))


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE DE SCORING (identique au script d'entraînement train_model.py)
# ─────────────────────────────────────────────────────────────────────────────
def predict(raw_fr: dict) -> float:
    """Dummifie, réindexe sur les colonnes du scaler, standardise puis prédit."""
    df = pd.DataFrame([raw_fr])
    df_dum = pd.get_dummies(df, dtype=int)
    df_dum = df_dum.reindex(columns=SCALER_COLS, fill_value=0)
    scaled = pd.DataFrame(scaler.transform(df_dum), columns=clean_cols)
    return float(model.predict_proba(scaled)[0][1])


# ─────────────────────────────────────────────────────────────────────────────
# GAUGE SVG SEMI-CIRCULAIRE
# ─────────────────────────────────────────────────────────────────────────────
def svg_gauge(prob: float, color: str) -> str:
    W, H   = 360, 218
    cx, cy = 180, 182
    R, r   = 138, 98

    def pt(deg, radius):
        rad = math.radians(deg)
        return cx + radius * math.cos(rad), cy - radius * math.sin(rad)

    def ring(d1, d2, fill, stroke):
        ox1, oy1 = pt(d1, R); ox2, oy2 = pt(d2, R)
        ix1, iy1 = pt(d1, r); ix2, iy2 = pt(d2, r)
        la = 1 if abs(d1 - d2) > 180 else 0
        d = (f"M{ox1:.2f} {oy1:.2f} A{R} {R} 0 {la} 0 {ox2:.2f} {oy2:.2f} "
             f"L{ix2:.2f} {iy2:.2f} A{r} {r} 0 {la} 1 {ix1:.2f} {iy1:.2f}Z")
        return f'<path d="{d}" fill="{fill}" stroke="{stroke}" stroke-width="1.8" stroke-linejoin="round"/>'

    d_g = 180 - 0.30 * 180   # 126°
    d_o = 180 - 0.50 * 180   # 90°
    d_n = 180 - prob * 180    # aiguille

    arcs = (ring(180, d_g, "#d5f5e3", "#27ae60")
          + ring(d_g, d_o,  "#fdf6d3", "#f1c40f")
          + ring(d_o, 0,    "#fadbd8", "#e74c3c"))

    ticks = ""
    for d in range(0, 181, 18):
        mx1, my1 = pt(180 - d, R + 3)
        mx2, my2 = pt(180 - d, R + 10)
        ticks += f'<line x1="{mx1:.1f}" y1="{my1:.1f}" x2="{mx2:.1f}" y2="{my2:.1f}" stroke="#bbb" stroke-width="1.5"/>'

    def zone_label(mid_deg, txt, col):
        lx, ly = pt(mid_deg, R + 22)
        return (f'<text x="{lx:.0f}" y="{ly:.0f}" '
                f'font-family="Inter,sans-serif" font-size="11.5" font-weight="700" '
                f'fill="{col}" text-anchor="middle">{txt}</text>')

    labels = (zone_label(153, t("risk_low"),  "#27ae60")
            + zone_label(108, t("risk_mod"),  "#b7950b")
            + zone_label(45,  t("risk_high"), "#e74c3c"))

    tick_labels = ""
    for deg, lbl, anchor in [(180,"0%","end"), (90,"50%","middle"), (0,"100%","start")]:
        tx, ty = pt(deg, R + 36)
        tick_labels += (f'<text x="{tx:.0f}" y="{ty:.0f}" '
                        f'font-family="Inter,sans-serif" font-size="9.5" '
                        f'fill="#aaa" text-anchor="{anchor}">{lbl}</text>')

    nx, ny   = pt(d_n, r - 6)
    nxb, nyb = pt(d_n + 90, 12)
    nxc, nyc = pt(d_n - 90, 12)
    needle = (f'<polygon points="{nx:.1f},{ny:.1f} {nxb:.1f},{nyb:.1f} {nxc:.1f},{nyc:.1f}" '
              f'fill="{color}" opacity="0.95"/>')
    hub    = (f'<circle cx="{cx}" cy="{cy}" r="11" fill="{color}"/>'
              f'<circle cx="{cx}" cy="{cy}" r="5.5" fill="white"/>')

    pct = (f'<text x="{cx}" y="{cy+36}" font-family="Inter,sans-serif" '
           f'font-size="34" font-weight="800" fill="{color}" text-anchor="middle">{prob:.1%}</text>')
    sub = (f'<text x="{cx}" y="{cy+54}" font-family="Inter,sans-serif" '
           f'font-size="10" fill="#999" text-anchor="middle">{t("prob_label")}</text>')

    return (f'<div style="text-align:center">'
            f'<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" '
            f'style="width:100%;max-width:390px;display:inline-block;">'
            f'{arcs}{ticks}{tick_labels}{labels}{needle}{hub}{pct}{sub}</svg></div>')


# ─────────────────────────────────────────────────────────────────────────────
# GRAPHIQUE IMPORTANCE
# ─────────────────────────────────────────────────────────────────────────────
VARIABLE_LABELS = {
    "Observance_4j":        {"fr": "Observance (4 jours)",         "en": "Adherence (4 days)"},
    "Region":               {"fr": "Région",                       "en": "Region"},
    "Religion":              {"fr": "Religion",                    "en": "Religion"},
    "Depenses_Mensuelles":  {"fr": "Dépenses mensuelles",          "en": "Monthly expenses"},
    "DSD_Recode":           {"fr": "Mode de dispensation",         "en": "Dispensation mode"},
    "Niveau_Etude":         {"fr": "Niveau d'étude",                "en": "Education level"},
    "Retesting":            {"fr": "Retesting VIH",                "en": "HIV retesting"},
    "Type_FOSA":            {"fr": "Type de FOSA",                 "en": "Facility type"},
    "Sexe":                 {"fr": "Sexe",                         "en": "Sex"},
    "Tranche_Age":          {"fr": "Tranche d'âge",                "en": "Age group"},
    "Statut_Matrimonial":   {"fr": "Statut matrimonial",           "en": "Marital status"},
    "Soutien_PEPFAR":       {"fr": "Soutien PEPFAR",               "en": "PEPFAR support"},
    "Soutien_Familial":     {"fr": "Soutien familial",             "en": "Family support"},
    "Traitement_Alternatif": {"fr": "Traitement alternatif",       "en": "Alternative treatment"},
}
MODALITY_EN = {
    "Oui": "Yes", "Non renseigné": "Not specified", "Mediocre": "Poor",
    "Modérée": "Moderate", "Standard": "Standard", "Public": "Public",
    "Masculin": "Male",
}


def get_feature_importances(mdl):
    # CalibratedClassifierCV (utilisé pour fiabiliser les probabilités) masque
    # feature_importances_/coef_ du modèle sous-jacent : on va les chercher
    # dans le premier estimateur calibré, moyennées sur les folds si possible.
    if hasattr(mdl, "calibrated_classifiers_"):
        import numpy as np
        estimateurs = [cc.estimator for cc in mdl.calibrated_classifiers_]
        if all(hasattr(e, "feature_importances_") for e in estimateurs):
            return np.mean([e.feature_importances_ for e in estimateurs], axis=0)
        if all(hasattr(e, "coef_") for e in estimateurs):
            coefs = [e.coef_[0] if e.coef_.ndim > 1 else e.coef_ for e in estimateurs]
            return np.abs(np.mean(coefs, axis=0))
        return None
    if hasattr(mdl, "feature_importances_"):
        return mdl.feature_importances_
    elif hasattr(mdl, "coef_"):
        import numpy as np
        coef = mdl.coef_
        if coef.ndim > 1:
            coef = coef[0]
        return np.abs(coef)
    return None


def get_label(var: str) -> str:
    for prefix, group in VARIABLE_LABELS.items():
        if var.startswith(prefix + "_"):
            modality = var[len(prefix) + 1:].replace("_", " ")
            if L == "en":
                modality = MODALITY_EN.get(modality, modality)
            return f"{group[L]} : {modality}"
    return var.replace("_", " ")


def draw_importance(top_n: int = 10) -> plt.Figure:
    imp  = get_feature_importances(model)
    feat = pd.DataFrame({"Variable": clean_cols, "Importance": imp})
    feat = feat.sort_values("Importance", ascending=False).head(top_n).sort_values("Importance")
    feat["Label"] = feat["Variable"].apply(get_label)

    n      = len(feat)
    COLORS = ["#2980b9","#2980b9","#27ae60","#27ae60","#f39c12",
               "#f39c12","#e67e22","#e74c3c","#c0392b","#922b21"]
    colors = COLORS[::-1][:n][::-1]

    fig, ax = plt.subplots(figsize=(7.2, 4.2))
    fig.patch.set_facecolor("#ffffff")
    ax.set_facecolor("#f7faff")
    bars = ax.barh(feat["Label"], feat["Importance"],
                   color=colors, height=0.60, edgecolor="#fff", linewidth=0.6)
    for bar, val in zip(bars, feat["Importance"]):
        ax.text(bar.get_width() + 0.0007, bar.get_y() + bar.get_height() / 2,
                f"{val:.4f}", va="center", ha="left", fontsize=8.5, color="#444", fontweight="500")
    ax.set_xlabel(t("imp_xlabel"), fontsize=8.5, color="#666")
    ax.set_title(t("imp_title").format(top_n), fontsize=10.5,
                 fontweight="700", color="#0b2d52", pad=12)
    ax.set_xlim(0, feat["Importance"].max() * 1.22)
    ax.tick_params(axis="y", labelsize=9, colors="#333", length=0)
    ax.tick_params(axis="x", labelsize=8, colors="#888")
    for sp in ["top", "right"]: ax.spines[sp].set_visible(False)
    ax.spines["left"].set_color("#e0e8f4")
    ax.spines["bottom"].set_color("#e0e8f4")
    ax.grid(axis="x", linestyle="--", alpha=0.4, color="#cce")
    plt.tight_layout(pad=0.9)
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# GÉNÉRATION PDF
# ─────────────────────────────────────────────────────────────────────────────
def safe(text: str) -> str:
    """Encode le texte en latin-1 pour fpdf2 (gère les accents français)."""
    return text.encode('latin-1', 'replace').decode('latin-1')


def generate_pdf(patient_vals: list, prob: float, niveau: str,
                 reco: str, lang: str, cnls_path: Path, issea_path: Path) -> bytes:
    """Génère une fiche PDF patient prête à imprimer."""

    if prob < 0.30:
        r, g, b       = 39, 174, 96
        bg_r, bg_g, bg_b = 213, 245, 227
    elif prob < 0.50:
        r, g, b       = 183, 149, 11
        bg_r, bg_g, bg_b = 255, 251, 230
    else:
        r, g, b       = 231, 76, 60
        bg_r, bg_g, bg_b = 253, 237, 236

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)

    # ── En-tête ──────────────────────────────────────────────────────────────
    pdf.set_fill_color(11, 45, 82)
    pdf.rect(0, 0, 210, 44, 'F')

    try:
        pdf.image(str(cnls_path),  x=8,  y=4,  h=36)
        pdf.image(str(issea_path), x=172, y=4, h=36)
    except Exception:
        pass

    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 13)
    pdf.set_xy(50, 7)
    pdf.cell(110, 7, safe('TARV-Score — Fiche de Scoring Clinique'), align='C', ln=True)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_xy(50, 17)
    lbl_sub = ("Outil de Detection du Risque d'Interruption au TARV chez les PVVIH au Cameroun"
               if lang == 'fr' else
               "Risk of ART Interruption Detection Tool among PLHIV in Cameroon")
    pdf.cell(110, 5, safe(lbl_sub), align='C', ln=True)
    pdf.set_xy(50, 25)
    pdf.cell(110, 5, safe('CNLS / GTC Cameroun | ISSEA-CEMAC 2025-2026'), align='C', ln=True)
    pdf.set_xy(50, 33)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.cell(110, 5, safe(f"Seuil : {SEUIL:.0%}"), align='C')

    # ── Date d'évaluation ────────────────────────────────────────────────────
    pdf.set_text_color(100, 100, 100)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_xy(10, 50)
    date_lbl = "Date d'evaluation" if lang == 'fr' else "Evaluation date"
    pdf.cell(0, 5, safe(f'{date_lbl} : {datetime.now().strftime("%d/%m/%Y  %H:%M")}'))

    # ── Score de risque ───────────────────────────────────────────────────────
    pdf.set_xy(10, 60)
    pdf.set_fill_color(bg_r, bg_g, bg_b)
    pdf.set_draw_color(r, g, b)
    pdf.set_line_width(0.8)
    pdf.rect(10, 60, 190, 30, 'FD')

    pdf.set_text_color(r, g, b)
    pdf.set_font('Helvetica', 'B', 28)
    pdf.set_xy(10, 64)
    pdf.cell(95, 12, f'{prob:.1%}', align='C')

    pdf.set_font('Helvetica', 'B', 18)
    pdf.set_xy(105, 64)
    risk_word = 'RISQUE' if lang == 'fr' else 'RISK'
    pdf.cell(95, 12, safe(f'{risk_word} {niveau}'), align='C')

    pdf.set_text_color(120, 120, 120)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_xy(10, 80)
    seuil_txt = (f'Seuil de classification : {SEUIL:.0%}'
                 if lang == 'fr' else
                 f'Classification threshold: {SEUIL:.0%}')
    pdf.cell(190, 5, safe(seuil_txt), align='C')

    # ── Profil du patient ─────────────────────────────────────────────────────
    pdf.set_text_color(11, 45, 82)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_xy(10, 96)
    pdf.cell(0, 7, safe('Profil du Patient' if lang == 'fr' else 'Patient Profile'), ln=True)

    pdf.set_fill_color(11, 45, 82)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_xy(10, 104)
    pdf.cell(85, 7, safe('Variable'), border=1, fill=True, align='C')
    pdf.cell(105, 7, safe('Valeur' if lang == 'fr' else 'Value'), border=1, fill=True, align='C', ln=True)

    var_labels = {
        'fr': ['Region', 'Type de FOSA', 'Mode DSD', 'Soutien PEPFAR',
               'Observance (4 jours)', 'Retesting VIH', 'Sexe', "Tranche d'age",
               'Statut matrimonial', 'Religion', "Niveau d'etude", 'Depenses mensuelles',
               'Soutien familial', 'Traitement alternatif'],
        'en': ['Region', 'Health Facility Type', 'DSD Mode', 'PEPFAR Support',
               'Adherence (4 days)', 'HIV Retesting', 'Sex', 'Age Group',
               'Marital Status', 'Religion', 'Education Level', 'Monthly Expenses',
               'Family support', 'Alternative treatment'],
    }
    pdf.set_line_width(0.3)
    for i, (var, val) in enumerate(zip(var_labels[lang], patient_vals)):
        pdf.set_fill_color(240, 245, 255) if i % 2 == 0 else pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(40, 40, 40)
        pdf.set_font('Helvetica', '', 8.5)
        pdf.cell(85, 6, safe(str(var)), border=1, fill=True)
        pdf.cell(105, 6, safe(str(val)), border=1, fill=True, ln=True)

    # ── Recommandation ────────────────────────────────────────────────────────
    pdf.ln(5)
    pdf.set_text_color(11, 45, 82)
    pdf.set_font('Helvetica', 'B', 10)
    reco_lbl = 'Recommandation clinique' if lang == 'fr' else 'Clinical Recommendation'
    pdf.cell(0, 7, safe(reco_lbl), ln=True)

    pdf.set_fill_color(bg_r, bg_g, bg_b)
    pdf.set_draw_color(r, g, b)
    pdf.set_line_width(0.8)
    pdf.set_text_color(50, 50, 50)
    pdf.set_font('Helvetica', '', 8.5)
    pdf.multi_cell(190, 5.5, safe(reco), border=1, fill=True)

    # ── Pied de page ──────────────────────────────────────────────────────────
    pdf.set_y(-15)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(160, 160, 160)
    disc = ("Outil d'aide a la decision. Ne remplace pas le jugement clinique du prestataire de sante."
            if lang == 'fr' else
            "Decision-support tool. Does not replace the clinical judgment of the healthcare provider.")
    pdf.cell(0, 5, safe(disc), align='C')

    return bytes(pdf.output())


# ─────────────────────────────────────────────────────────────────────────────
# FONCTIONS IMPORT BATCH
# ─────────────────────────────────────────────────────────────────────────────
COLS_REQUIS = [
    "Region", "Type_FOSA", "DSD_Recode", "Soutien_PEPFAR",
    "Observance_4j", "Retesting", "Sexe", "Tranche_Age",
    "Statut_Matrimonial", "Religion", "Niveau_Etude", "Depenses_Mensuelles",
    "Soutien_Familial", "Traitement_Alternatif",
]

VALEURS_VALIDES = {
    "Region":              T["region_opts"]["fr"],
    "Type_FOSA":           T["type_fosa_opts"]["fr"],
    "DSD_Recode":          T["dsd_opts"]["fr"],
    "Soutien_PEPFAR":      ["Oui", "Non"],
    "Observance_4j":       T["observance_opts"]["fr"],
    "Retesting":           ["Oui", "Non"],
    "Sexe":                ["Féminin", "Masculin"],
    "Tranche_Age":         T["tranche_age_opts"]["fr"],
    "Statut_Matrimonial":  T["statut_mat_opts"]["fr"],
    "Religion":            T["religion_opts"]["fr"],
    "Niveau_Etude":        T["niveau_etude_opts"]["fr"],
    "Depenses_Mensuelles": T["depenses_opts"]["fr"],
    "Soutien_Familial":    ["Oui", "Non"],
    "Traitement_Alternatif": ["Oui", "Non"],
}


def generate_template() -> bytes:
    """Génère un fichier Excel modèle avec les colonnes et valeurs valides."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Données patients"

    header_fill  = PatternFill("solid", fgColor="0B2D52")
    header_font  = Font(color="FFFFFF", bold=True, size=10)
    border_side  = Side(style="thin", color="CCCCCC")
    cell_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers = ["ID_Patient"] + COLS_REQUIS
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(c)].width = 22

    exemple = ["P001", "Centre", "Public", "Standard", "Oui",
               "Bonne", "Non", "Féminin", "25 à 49 Ans",
               "Marié(e) en monogamie", "Catholique", "Jamais fréquenté",
               "Moins de 5 000", "Oui", "Non"]
    for c, val in enumerate(exemple, 1):
        cell = ws.cell(row=2, column=c, value=val)
        cell.fill = PatternFill("solid", fgColor="EEF5FF")
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 32

    ws2 = wb.create_sheet("Valeurs valides")
    ws2.cell(1, 1, "Variable").font = Font(bold=True, color="FFFFFF", size=10)
    ws2.cell(1, 1).fill = header_fill
    ws2.cell(1, 2, "Valeurs acceptées").font = Font(bold=True, color="FFFFFF", size=10)
    ws2.cell(1, 2).fill = header_fill
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 80

    row = 2
    for var, vals in VALEURS_VALIDES.items():
        ws2.cell(row, 1, var).font = Font(bold=True, color="0B2D52")
        ws2.cell(row, 2, " | ".join(vals))
        ws2.cell(row, 2).alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[row].height = 20
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def normalize_fosa(val: str) -> str:
    mapping = {"Privé laïc": "Privé laic", "Prive laic": "Privé laic",
               "Prive confessionnel": "Privé confessionnel"}
    return mapping.get(str(val).strip(), str(val).strip())


def normalize_observance(val: str) -> str:
    mapping = {"Médiocre": "Mediocre", "Mediocre": "Mediocre",
               "Modérée": "Modérée", "Bonne": "Bonne",
               "Poor": "Mediocre", "Moderate": "Modérée", "Good": "Bonne"}
    return mapping.get(str(val).strip(), str(val).strip())


def score_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Score tous les patients d'un DataFrame et retourne df enrichi."""
    results = []
    for _, row in df.iterrows():
        try:
            raw = {
                "Region":              str(row.get("Region", "")).strip(),
                "Type_FOSA":           normalize_fosa(row.get("Type_FOSA", "")),
                "DSD_Recode":          str(row.get("DSD_Recode", "")).strip(),
                "Soutien_PEPFAR":      str(row.get("Soutien_PEPFAR", "")).strip(),
                "Observance_4j":       normalize_observance(row.get("Observance_4j", "")),
                "Retesting":           str(row.get("Retesting", "")).strip(),
                "Sexe":                str(row.get("Sexe", "")).strip(),
                "Tranche_Age":         str(row.get("Tranche_Age", "")).strip(),
                "Statut_Matrimonial":  str(row.get("Statut_Matrimonial", "")).strip(),
                "Religion":            str(row.get("Religion", "")).strip(),
                "Niveau_Etude":        str(row.get("Niveau_Etude", "")).strip(),
                "Depenses_Mensuelles": str(row.get("Depenses_Mensuelles", "")).strip(),
                "Soutien_Familial":    str(row.get("Soutien_Familial", "")).strip(),
                "Traitement_Alternatif": str(row.get("Traitement_Alternatif", "")).strip(),
            }
            prob = predict(raw)
            if prob < 0.30:
                niv = "Faible / Low"
            elif prob < SEUIL:
                niv = "Modéré / Moderate"
            else:
                niv = "Élevé / High"
            results.append({"Probabilité (%)": round(prob * 100, 1), "Niveau de risque": niv, "Erreur": ""})
        except Exception as e:
            results.append({"Probabilité (%)": None, "Niveau de risque": "—", "Erreur": str(e)})

    return pd.concat([df.reset_index(drop=True), pd.DataFrame(results)], axis=1)


DASHBOARD_VARS = {
    "Region":        {"fr": "Région",     "en": "Region"},
    "Type_FOSA":     {"fr": "Type de FOSA", "en": "Facility type"},
    "Observance_4j": {"fr": "Observance",  "en": "Adherence"},
    "Sexe":          {"fr": "Sexe",        "en": "Sex"},
}


def draw_dashboard(df_res: pd.DataFrame, lang: str) -> plt.Figure:
    """Répartition des niveaux de risque et taux de risque élevé par variable clé."""
    niveau_order = ["Faible / Low", "Modéré / Moderate", "Élevé / High"]
    niveau_lbl   = {"fr": ["Faible", "Modéré", "Élevé"], "en": ["Low", "Moderate", "High"]}
    colors_risk  = ["#27ae60", "#f1c40f", "#e74c3c"]

    fig, axes = plt.subplots(2, 2, figsize=(11, 7.5))
    fig.patch.set_facecolor("#ffffff")

    ax = axes[0, 0]
    counts = df_res["Niveau de risque"].value_counts().reindex(niveau_order).fillna(0)
    ax.pie(counts, labels=niveau_lbl[lang], autopct=lambda p: f"{p:.0f}%" if p > 0 else "",
           colors=colors_risk, wedgeprops={"edgecolor": "white", "linewidth": 1.5},
           textprops={"fontsize": 9})
    ax.set_title("Répartition des niveaux de risque" if lang == "fr" else "Risk level distribution",
                 fontsize=10.5, fontweight="700", color="#0b2d52", pad=10)

    axes_bar = [axes[0, 1], axes[1, 0], axes[1, 1]]
    for ax, var in zip(axes_bar, ["Region", "Type_FOSA", "Observance_4j"]):
        ax.set_facecolor("#f7faff")
        if var not in df_res.columns:
            ax.axis("off")
            continue
        grp = (df_res.groupby(var)["Niveau de risque"]
               .apply(lambda s: (s == "Élevé / High").mean() * 100)
               .sort_values())
        ax.barh(grp.index, grp.values, color="#e74c3c", height=0.6, edgecolor="#fff", linewidth=0.5)
        for y, v in enumerate(grp.values):
            ax.text(v + 1, y, f"{v:.0f}%", va="center", fontsize=8, color="#444")
        label = DASHBOARD_VARS.get(var, {}).get(lang, var)
        titre = f"% à risque élevé par {label}" if lang == "fr" else f"% high risk by {label}"
        ax.set_title(titre, fontsize=10, fontweight="700", color="#0b2d52", pad=8)
        ax.set_xlim(0, max(grp.values.max() * 1.25, 10) if len(grp) else 10)
        ax.tick_params(labelsize=8.5, colors="#333")
        for sp in ["top", "right"]:
            ax.spines[sp].set_visible(False)
        ax.spines["left"].set_color("#e0e8f4")
        ax.spines["bottom"].set_color("#e0e8f4")

    plt.tight_layout(pad=1.4)
    return fig


def generate_batch_pdf(df_res: pd.DataFrame, cnls_path: Path, issea_path: Path) -> bytes:
    """PDF récapitulatif de tous les patients scorés."""
    pdf = FPDF(orientation='L')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=12)

    pdf.set_fill_color(11, 45, 82)
    pdf.rect(0, 0, 297, 30, 'F')
    try:
        pdf.image(str(cnls_path),  x=4,  y=2, h=26)
        pdf.image(str(issea_path), x=263, y=2, h=26)
    except Exception:
        pass
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_xy(40, 5)
    pdf.cell(217, 8, safe('TARV-Score — Rapport de Scoring par Lot'), align='C', ln=True)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_xy(40, 16)
    pdf.cell(217, 5, safe(f'CNLS / GTC Cameroun | {datetime.now().strftime("%d/%m/%Y %H:%M")} | {len(df_res)} patients'), align='C')

    pdf.set_xy(10, 36)
    pdf.set_text_color(11, 45, 82)
    pdf.set_font('Helvetica', 'B', 10)
    n_tot  = len(df_res)
    n_low  = (df_res["Niveau de risque"].str.startswith("Faible")).sum()
    n_mod  = (df_res["Niveau de risque"].str.startswith("Mod")).sum()
    n_high = (df_res["Niveau de risque"].str.startswith("El")).sum()
    pdf.cell(0, 6, safe(f'Resume : {n_tot} patients | Faible : {n_low} | Modere : {n_mod} | Eleve : {n_high}'), ln=True)

    pdf.ln(2)
    cols_show = (["ID_Patient"] if "ID_Patient" in df_res.columns else []) + \
                COLS_REQUIS + ["Probabilité (%)", "Niveau de risque"]
    col_widths = [round(277 / len(cols_show), 1)] * len(cols_show)
    if "ID_Patient" in cols_show:
        col_widths[0] = 14

    pdf.set_fill_color(11, 45, 82)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 6.5)
    for col, w in zip(cols_show, col_widths):
        pdf.cell(w, 7, safe(col[:16]), border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('Helvetica', '', 6.5)
    for i, (_, row) in enumerate(df_res.iterrows()):
        niv = str(row.get("Niveau de risque", ""))
        if niv.startswith("El"):
            pdf.set_fill_color(253, 237, 236)
            pdf.set_text_color(180, 40, 30)
        elif niv.startswith("Mod"):
            pdf.set_fill_color(255, 251, 230)
            pdf.set_text_color(140, 115, 9)
        else:
            pdf.set_fill_color(213, 245, 227) if i % 2 == 0 else pdf.set_fill_color(240, 255, 245)
            pdf.set_text_color(30, 100, 60)

        for col, w in zip(cols_show, col_widths):
            val = str(row.get(col, ""))[:16]
            pdf.cell(w, 6, safe(val), border=1, fill=True, align='C')
        pdf.ln()

    pdf.set_y(-12)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(160, 160, 160)
    pdf.cell(0, 5, safe("Outil d'aide a la decision — Ne remplace pas le jugement clinique | TARV-Score | ISSEA-CEMAC 2025-2026"), align='C')

    return bytes(pdf.output())


# ─────────────────────────────────────────────────────────────────────────────
# ONGLETS PRINCIPAUX
# ─────────────────────────────────────────────────────────────────────────────
tab_lbl1 = "📝 Saisie individuelle" if L == "fr" else "📝 Individual entry"
tab_lbl2 = "📂 Import fichier (Excel / CSV)" if L == "fr" else "📂 File import (Excel / CSV)"
tab_lbl3 = "🔬 Diagnostic & Vérification" if L == "fr" else "🔬 Diagnostic & Verification"
tab1, tab2, tab3 = st.tabs([tab_lbl1, tab_lbl2, tab_lbl3])

with tab1:

    st.markdown(f"""
<div style="background:#fff;border-radius:14px;padding:18px 22px 6px 22px;
     margin-bottom:4px;box-shadow:0 2px 10px rgba(0,0,0,0.06);border-top:4px solid #1a5e8a;">
  <div style="font-size:1.02em;font-weight:700;color:#0b2d52;">{t("form_intro")}</div>
  <div style="font-size:0.84em;color:#777;margin-top:3px;">{t("form_sub")}</div>
</div>
""", unsafe_allow_html=True)

    with st.form("patient_form", clear_on_submit=False):
        col1, col2, col3 = st.columns(3, gap="medium")

        with col1:
            st.markdown(
                f'<div style="font-size:0.82em;font-weight:700;color:#1a5e8a;'
                f'text-transform:uppercase;letter-spacing:1px;padding:8px 0 6px 0;'
                f'border-bottom:2px solid #d0e4f7;margin-bottom:12px;">'
                f'{t("sec_soins")}</div>', unsafe_allow_html=True)
            region    = st.selectbox(t("region"),    T["region_opts"][L],    key=f"{L}_region")
            type_fosa = st.selectbox(t("type_fosa"), T["type_fosa_opts"][L], key=f"{L}_fosa")
            dsd       = st.selectbox(t("dsd"),       T["dsd_opts"][L],       key=f"{L}_dsd")
            pepfar    = st.radio(t("pepfar"), [t("non"), t("oui")], horizontal=True, key=f"{L}_pepfar")
            soutien_familial = st.radio(t("soutien_familial"), [t("non"), t("oui")], horizontal=True, key=f"{L}_soutien_fam")

        with col2:
            st.markdown(
                f'<div style="font-size:0.82em;font-weight:700;color:#0d7a5c;'
                f'text-transform:uppercase;letter-spacing:1px;padding:8px 0 6px 0;'
                f'border-bottom:2px solid #c3e8da;margin-bottom:12px;">'
                f'{t("sec_thera")}</div>', unsafe_allow_html=True)
            observance = st.selectbox(t("observance"), T["observance_opts"][L], key=f"{L}_observance")
            retesting  = st.radio(t("retesting"), [t("non"), t("oui")], horizontal=True, key=f"{L}_retesting")
            traitement_alt = st.radio(t("traitement_alt"), [t("non"), t("oui")], horizontal=True, key=f"{L}_traitement_alt")
            sexe       = st.radio(t("sexe"), [t("feminin"), t("masculin")], horizontal=True, key=f"{L}_sexe")
            tranche_age = st.selectbox(t("tranche_age"), T["tranche_age_opts"][L], key=f"{L}_age")

        with col3:
            st.markdown(
                f'<div style="font-size:0.82em;font-weight:700;color:#6c3483;'
                f'text-transform:uppercase;letter-spacing:1px;padding:8px 0 6px 0;'
                f'border-bottom:2px solid #e8d5f5;margin-bottom:12px;">'
                f'{t("sec_socio")}</div>', unsafe_allow_html=True)
            statut_mat   = st.selectbox(t("statut_mat"),   T["statut_mat_opts"][L],   key=f"{L}_statut")
            religion     = st.selectbox(t("religion"),     T["religion_opts"][L],     key=f"{L}_religion")
            niveau_etude = st.selectbox(t("niveau_etude"), T["niveau_etude_opts"][L], key=f"{L}_etude")
            depenses     = st.selectbox(t("depenses"),     T["depenses_opts"][L],     key=f"{L}_depenses")

        st.markdown("<br>", unsafe_allow_html=True)
        submitted = st.form_submit_button(t("btn_calc"), use_container_width=True, type="primary")

    # ─────────────────────────────────────────────────────────────────────────
    # RÉSULTATS
    # ─────────────────────────────────────────────────────────────────────────
    if submitted:
        region_fr    = REGION_EN2FR.get(region, region)          if L == "en" else region
        type_fosa_fr = FOSA_INTERNAL[type_fosa]
        dsd_fr       = DSD_EN2FR.get(dsd, dsd)                    if L == "en" else dsd
        pepfar_fr    = "Oui" if pepfar == t("oui") else "Non"
        soutien_fam_fr = "Oui" if soutien_familial == t("oui") else "Non"
        observ_fr    = OBSERVANCE_INTERNAL[observance]
        retesting_fr = "Oui" if retesting == t("oui") else "Non"
        traitement_alt_fr = "Oui" if traitement_alt == t("oui") else "Non"
        sexe_fr      = "Masculin" if sexe == t("masculin") else "Féminin"
        tranche_fr   = TRANCHE_EN2FR.get(tranche_age, tranche_age) if L == "en" else tranche_age
        statut_fr    = STATUT_EN2FR.get(statut_mat, statut_mat)   if L == "en" else statut_mat
        religion_fr  = RELIGION_EN2FR.get(religion, religion)     if L == "en" else religion
        etude_fr     = ETUDE_EN2FR.get(niveau_etude, niveau_etude) if L == "en" else niveau_etude
        depenses_fr  = DEPENSES_EN2FR.get(depenses, depenses)     if L == "en" else depenses

        raw_fr = {
            "Region":              region_fr,
            "Type_FOSA":           type_fosa_fr,
            "DSD_Recode":          dsd_fr,
            "Soutien_PEPFAR":      pepfar_fr,
            "Observance_4j":       observ_fr,
            "Retesting":           retesting_fr,
            "Sexe":                sexe_fr,
            "Tranche_Age":         tranche_fr,
            "Statut_Matrimonial":  statut_fr,
            "Religion":            religion_fr,
            "Niveau_Etude":        etude_fr,
            "Depenses_Mensuelles": depenses_fr,
            "Soutien_Familial":    soutien_fam_fr,
            "Traitement_Alternatif": traitement_alt_fr,
        }

        with st.spinner(t("spinner")):
            prob = predict(raw_fr)

        if prob < 0.30:
            niveau, color        = t("risk_low_lbl"),  "#27ae60"
            emoji_r              = "🟢"
            bg_card, bd_card     = "#eafaf1", "#27ae60"
            bg_reco, bd_reco, tc = "#eafaf1", "#27ae60", "#1e5f3a"
            reco = t("reco_low")
        elif prob < SEUIL:
            niveau, color        = t("risk_mod_lbl"),  "#b7950b"
            emoji_r              = "🟡"
            bg_card, bd_card     = "#fffbe6", "#f1c40f"
            bg_reco, bd_reco, tc = "#fffbe6", "#f1c40f", "#7a6200"
            reco = t("reco_mod")
        else:
            niveau, color        = t("risk_high_lbl"), "#e74c3c"
            emoji_r              = "🔴"
            bg_card, bd_card     = "#fdedec", "#e74c3c"
            bg_reco, bd_reco, tc = "#fdedec", "#e74c3c", "#7b241c"
            reco = t("reco_high")

        risk_word = "RISQUE" if L == "fr" else "RISK"

        st.markdown(
            f'<div class="res-divider"><hr><span>{t("res_divider")}</span><hr></div>',
            unsafe_allow_html=True,
        )

        col_g, col_r = st.columns([1, 1.45], gap="large")

        with col_g:
            st.markdown(svg_gauge(prob, color), unsafe_allow_html=True)
            st.markdown(f"""
            <div style="background:{bg_card};border-radius:14px;border:2px solid {bd_card};
                 padding:18px 20px;text-align:center;margin-top:12px;
                 box-shadow:0 4px 16px {bd_card}33;">
                <div style="font-size:0.78em;font-weight:800;letter-spacing:2.5px;
                     text-transform:uppercase;color:{color};margin-bottom:4px;">
                    {emoji_r} &nbsp; {risk_word} {niveau}
                </div>
                <div style="font-size:3em;font-weight:900;color:{color};line-height:1.1;">
                    {prob:.1%}
                </div>
                <div style="font-size:0.8em;color:#777;margin-top:4px;">
                    {t("seuil_txt")} : {SEUIL:.0%}
                </div>
            </div>
            <div style="background:{bg_reco};border-left:5px solid {bd_reco};
                 border-radius:0 10px 10px 0;padding:14px 16px;
                 margin-top:10px;font-size:0.88em;line-height:1.6;color:{tc};">
                <strong>{t("reco_lbl")}</strong><br>{reco}
            </div>
            """, unsafe_allow_html=True)

        with col_r:
            if get_feature_importances(model) is not None:
                st.markdown(
                    f'<div style="font-size:0.83em;color:#777;margin-bottom:6px;">{t("imp_sub")}</div>',
                    unsafe_allow_html=True,
                )
                st.pyplot(draw_importance(10), use_container_width=True)
                st.markdown(
                    f'<div style="font-size:0.78em;color:#888;margin-top:2px;text-align:center;">'
                    f'{t("imp_legend")}</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.info("📊 Ce modèle ne fournit pas d'importances de variables." if L=="fr"
                        else "📊 This model does not provide feature importances.")

            with st.expander(t("recap_title")):
                var_labels = {
                    "fr": ["Région","Type de FOSA","Mode DSD","Soutien PEPFAR",
                           "Observance (4j)","Retesting VIH","Sexe","Tranche d'âge",
                           "Statut matrimonial","Religion","Niveau d'étude","Dépenses mensuelles",
                           "Soutien familial","Traitement alternatif"],
                    "en": ["Region","Health Facility Type","DSD Mode","PEPFAR Support",
                           "Adherence (4 days)","HIV Retesting","Sex","Age Group",
                           "Marital Status","Religion","Education Level","Monthly Expenses",
                           "Family support","Alternative treatment"],
                }
                recap = pd.DataFrame({
                    t("recap_var"): var_labels[L],
                    t("recap_val"): [region, type_fosa, dsd, pepfar,
                                     observance, retesting, sexe, tranche_age,
                                     statut_mat, religion, niveau_etude, depenses,
                                     soutien_familial, traitement_alt],
                })
                st.dataframe(recap, use_container_width=True, hide_index=True)

        # ── Bouton export PDF ─────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            f'<div style="background:#fff;border-radius:12px;padding:16px 20px;'
            f'border:2px dashed #1a5e8a;text-align:center;margin-top:8px;">'
            f'<div style="font-size:0.9em;font-weight:600;color:#0b2d52;margin-bottom:10px;">'
            f'{"📄 Télécharger la fiche patient (PDF imprimable)" if L=="fr" else "📄 Download patient report (printable PDF)"}'
            f'</div></div>',
            unsafe_allow_html=True,
        )
        patient_vals = [region, type_fosa, dsd, pepfar, observance, retesting,
                        sexe, tranche_age, statut_mat, religion, niveau_etude, depenses,
                        soutien_familial, traitement_alt]
        try:
            pdf_bytes = generate_pdf(
                patient_vals=patient_vals,
                prob=prob,
                niveau=niveau,
                reco=reco,
                lang=L,
                cnls_path=ASSETS / "cnls_logo.png",
                issea_path=ASSETS / "issea_logo.png",
            )
            btn_lbl = "⬇️ Télécharger le PDF" if L == "fr" else "⬇️ Download PDF"
            st.download_button(
                label=btn_lbl,
                data=pdf_bytes,
                file_name=f"fiche_TARV_Score_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"PDF non disponible : {e}")

with tab2:

    intro_txt = ("Uploadez un fichier Excel ou CSV contenant les données de plusieurs patients. "
                 "L'application les scorera automatiquement et vous pourrez télécharger les résultats."
                 if L == "fr" else
                 "Upload an Excel or CSV file with multiple patient records. "
                 "The app will score them automatically and you can download the results.")
    st.markdown(f"""
    <div style="background:#fff;border-radius:14px;padding:16px 22px;
         margin-bottom:16px;box-shadow:0 2px 10px rgba(0,0,0,0.06);border-top:4px solid #0d7a5c;">
      <div style="font-size:1em;font-weight:700;color:#0b2d52;">
          {"📂 Scoring par lot — Import de fichier" if L=="fr" else "📂 Batch Scoring — File Import"}
      </div>
      <div style="font-size:0.84em;color:#777;margin-top:4px;">{intro_txt}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown(
        f'<div style="font-weight:600;color:#0b2d52;margin-bottom:6px;">'
        f'{"📋 Étape 1 — Téléchargez le modèle Excel à remplir" if L=="fr" else "📋 Step 1 — Download the Excel template to fill in"}'
        f'</div>', unsafe_allow_html=True)

    try:
        tpl_bytes = generate_template()
        st.download_button(
            label="⬇️ Télécharger le modèle Excel" if L == "fr" else "⬇️ Download Excel template",
            data=tpl_bytes,
            file_name="modele_TARV_Score.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )
    except Exception as e:
        st.error(f"Erreur génération modèle : {e}")

    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown(
        f'<div style="font-weight:600;color:#0b2d52;margin-bottom:6px;">'
        f'{"📤 Étape 2 — Uploadez votre fichier rempli" if L=="fr" else "📤 Step 2 — Upload your filled file"}'
        f'</div>', unsafe_allow_html=True)

    upload_label = "Choisir un fichier Excel ou CSV" if L == "fr" else "Choose an Excel or CSV file"
    uploaded = st.file_uploader(upload_label, type=["xlsx", "xls", "csv"], key="batch_upload")

    if uploaded is not None:
        try:
            if uploaded.name.lower().endswith(".csv"):
                df_up = pd.read_csv(uploaded, sep=None, engine="python")
            else:
                df_up = pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"Impossible de lire le fichier : {e}")
            df_up = None

        if df_up is not None:
            missing = [c for c in COLS_REQUIS if c not in df_up.columns]
            if missing:
                st.error(
                    ("❌ Colonnes manquantes dans le fichier : **" + "**, **".join(missing) + "**\n\n"
                     "Utilisez le modèle Excel fourni à l'étape 1."
                     if L == "fr" else
                     "❌ Missing columns in file: **" + "**, **".join(missing) + "**\n\n"
                     "Please use the Excel template from step 1.")
                )
            else:
                st.success(f"✅ {'Fichier valide' if L=='fr' else 'Valid file'} — {len(df_up)} {'patients détectés' if L=='fr' else 'patients detected'}")

                with st.expander("👁️ Aperçu du fichier" if L == "fr" else "👁️ File preview"):
                    st.dataframe(df_up.head(5), use_container_width=True)

                btn_score = "🔍 Scorer tous les patients" if L == "fr" else "🔍 Score all patients"
                if st.button(btn_score, type="primary", use_container_width=True, key="batch_score_btn"):
                    with st.spinner("Calcul en cours…" if L == "fr" else "Computing…"):
                        df_res = score_dataframe(df_up)

                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown(
                        f'<div style="font-weight:700;font-size:1em;color:#0b2d52;margin-bottom:8px;">'
                        f'{"📊 Résultats du scoring" if L=="fr" else "📊 Scoring Results"}'
                        f'</div>', unsafe_allow_html=True)

                    def color_row(row):
                        niv = str(row.get("Niveau de risque", ""))
                        if niv.startswith("El"):
                            return ["background-color:#fdedec"] * len(row)
                        elif niv.startswith("Mod"):
                            return ["background-color:#fffbe6"] * len(row)
                        else:
                            return ["background-color:#eafaf1"] * len(row)

                    st.dataframe(
                        df_res.style.apply(color_row, axis=1),
                        use_container_width=True,
                        height=min(400, 40 + len(df_res) * 35),
                    )

                    n_tot  = len(df_res)
                    n_low  = (df_res["Niveau de risque"].str.startswith("Faible")).sum()
                    n_mod  = (df_res["Niveau de risque"].str.startswith("Mod")).sum()
                    n_high = (df_res["Niveau de risque"].str.startswith("El")).sum()
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Total", n_tot)
                    c2.metric("🟢 Faible / Low",   n_low)
                    c3.metric("🟡 Modéré / Mod",   n_mod)
                    c4.metric("🔴 Élevé / High",   n_high)

                    st.markdown("<br>", unsafe_allow_html=True)
                    dash_titre = "📊 Tableau de bord descriptif" if L == "fr" else "📊 Descriptive dashboard"
                    st.markdown(f"#### {dash_titre}")
                    try:
                        st.pyplot(draw_dashboard(df_res, L), use_container_width=True)
                    except Exception as e:
                        st.warning(f"Tableau de bord non disponible : {e}" if L == "fr"
                                   else f"Dashboard not available: {e}")

                    st.markdown("<br>", unsafe_allow_html=True)
                    dl1, dl2 = st.columns(2)

                    with dl1:
                        import io as _io
                        buf = _io.BytesIO()
                        df_res.to_excel(buf, index=False, engine="openpyxl")
                        buf.seek(0)
                        st.download_button(
                            label="⬇️ Télécharger Excel (avec scores)" if L == "fr" else "⬇️ Download Excel (with scores)",
                            data=buf.read(),
                            file_name=f"TARV_scores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                    with dl2:
                        try:
                            pdf_batch = generate_batch_pdf(
                                df_res,
                                cnls_path=ASSETS / "cnls_logo.png",
                                issea_path=ASSETS / "issea_logo.png",
                            )
                            st.download_button(
                                label="⬇️ Télécharger PDF récapitulatif" if L == "fr" else "⬇️ Download PDF summary",
                                data=pdf_batch,
                                file_name=f"TARV_rapport_lot_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                            )
                        except Exception as e:
                            st.warning(f"PDF non disponible : {e}")

with tab3:
    diag_title = "🔬 Vérification du pipeline de scoring" if L == "fr" else "🔬 Scoring pipeline verification"
    diag_sub = (
        "Cet onglet permet de vérifier que le modèle score correctement en testant des profils de référence connus. "
        "Les probabilités affichées reflètent ce que le modèle a appris des données d'entraînement."
        if L == "fr" else
        "This tab verifies that the model scores correctly by testing known reference profiles. "
        "Displayed probabilities reflect what the model learned from training data."
    )
    st.markdown(f"""
    <div style="background:#fff;border-radius:14px;padding:18px 22px;
         margin-bottom:16px;box-shadow:0 2px 10px rgba(0,0,0,0.06);border-top:4px solid #6c3483;">
      <div style="font-size:1em;font-weight:700;color:#0b2d52;">{diag_title}</div>
      <div style="font-size:0.84em;color:#777;margin-top:4px;">{diag_sub}</div>
    </div>
    """, unsafe_allow_html=True)

    note_titre = "ℹ️ Comprendre les résultats" if L == "fr" else "ℹ️ Understanding the results"
    note_corps = {
        "fr": """
**Comment lire le score ?**

L'outil a été validé sur les 14 caractéristiques les plus associées à l'interruption du TARV, avec une méthode de sélection statistique robuste, ne privilégiant aucun critère unique.

**Les catégories de référence** (Centre, FOSA publique, Standard, observance bonne...) sont les profils les plus **communs dans les données**. Le score de chaque patient est calculé par comparaison à ce profil de base.

> **Conseil clinique** : utilisez le score comme **signal de risque relatif** (comparer des patients entre eux) plutôt qu'en valeur absolue.
        """,
        "en": """
**How to read the score?**

The tool was validated on the 14 characteristics most associated with ART interruption, using a robust statistical selection method that does not favour any single criterion.

**Reference categories** (Centre, public facility, Standard, good adherence...) are the **most common profiles in the data**. Each patient's score is computed by comparison to this baseline profile.

> **Clinical advice**: use the score as a **relative risk signal** (compare patients to each other) rather than in absolute value.
        """
    }
    with st.expander(note_titre, expanded=True):
        st.markdown(note_corps[L])

    st.divider()

    test_titre = "🧪 Profils de test de référence" if L == "fr" else "🧪 Reference test profiles"
    st.markdown(f"#### {test_titre}")
    test_sub = (
        "Appuyez sur le bouton pour scorer 4 profils de référence et vérifier le comportement du modèle."
        if L == "fr" else
        "Click the button to score 4 reference profiles and verify the model's behavior."
    )
    st.markdown(f'<div style="font-size:0.85em;color:#666;margin-bottom:12px;">{test_sub}</div>',
                unsafe_allow_html=True)

    PROFILS_TEST = [
        {
            "nom_fr": "Profil 1 — Toutes catégories de référence",
            "nom_en": "Profile 1 — All reference categories",
            "risque_attendu_fr": "Risque de référence (baseline du modèle)",
            "risque_attendu_en": "Reference risk (model baseline)",
            "raw": {
                "Region": "Centre", "Type_FOSA": "Public", "DSD_Recode": "Standard",
                "Soutien_PEPFAR": "Oui", "Observance_4j": "Bonne", "Retesting": "Non",
                "Sexe": "Féminin", "Tranche_Age": "25 à 49 Ans",
                "Statut_Matrimonial": "Marié(e) en monogamie",
                "Religion": "Catholique", "Niveau_Etude": "Jamais fréquenté",
                "Depenses_Mensuelles": "Moins de 5 000",
                "Soutien_Familial": "Oui", "Traitement_Alternatif": "Non",
            },
        },
        {
            "nom_fr": "Profil 2 — Observance médiocre, jeune, célibataire",
            "nom_en": "Profile 2 — Poor adherence, young, single",
            "risque_attendu_fr": "Risque élevé attendu",
            "risque_attendu_en": "Expected high risk",
            "raw": {
                "Region": "Extrême-Nord", "Type_FOSA": "Privé laic", "DSD_Recode": "DSD avec décalage RDV",
                "Soutien_PEPFAR": "Non", "Observance_4j": "Mediocre", "Retesting": "Oui",
                "Sexe": "Masculin", "Tranche_Age": "18 à 20 Ans",
                "Statut_Matrimonial": "Célibataire",
                "Religion": "Musulman", "Niveau_Etude": "Jamais fréquenté",
                "Depenses_Mensuelles": "25 000 et plus",
                "Soutien_Familial": "Non", "Traitement_Alternatif": "Oui",
            },
        },
        {
            "nom_fr": "Profil 3 — Observance modérée, profil intermédiaire",
            "nom_en": "Profile 3 — Moderate adherence, intermediate profile",
            "risque_attendu_fr": "Risque modéré attendu",
            "risque_attendu_en": "Expected moderate risk",
            "raw": {
                "Region": "Ouest", "Type_FOSA": "Privé confessionnel", "DSD_Recode": "DSD sans décalage",
                "Soutien_PEPFAR": "Oui", "Observance_4j": "Modérée", "Retesting": "Non",
                "Sexe": "Masculin", "Tranche_Age": "50 ans et plus",
                "Statut_Matrimonial": "Veuf (ve)",
                "Religion": "Protestant", "Niveau_Etude": "Secondaire Premier Cycle",
                "Depenses_Mensuelles": "[10 000 - 25 000[",
                "Soutien_Familial": "Oui", "Traitement_Alternatif": "Non",
            },
        },
        {
            "nom_fr": "Profil 4 — Bonne observance, profil favorisé",
            "nom_en": "Profile 4 — Good adherence, favourable profile",
            "risque_attendu_fr": "Risque faible attendu",
            "risque_attendu_en": "Expected low risk",
            "raw": {
                "Region": "Littoral", "Type_FOSA": "Public", "DSD_Recode": "Standard",
                "Soutien_PEPFAR": "Oui", "Observance_4j": "Bonne", "Retesting": "Non",
                "Sexe": "Féminin", "Tranche_Age": "25 à 49 Ans",
                "Statut_Matrimonial": "Marié(e) en monogamie",
                "Religion": "Catholique", "Niveau_Etude": "Supérieur",
                "Depenses_Mensuelles": "Moins de 5 000",
                "Soutien_Familial": "Oui", "Traitement_Alternatif": "Non",
            },
        },
    ]

    btn_diag = "▶ Lancer le diagnostic" if L == "fr" else "▶ Run diagnostic"
    if st.button(btn_diag, type="primary", key="btn_diag"):
        st.markdown("<br>", unsafe_allow_html=True)
        rows = []
        for p in PROFILS_TEST:
            nom = p[f"nom_{L}"]
            attendu = p[f"risque_attendu_{L}"]
            try:
                p_val = predict(p["raw"])
                score_str = f"{p_val:.1%}"
            except Exception as e:
                score_str = f"Err: {e}"
            rows.append({
                "Profil" if L == "fr" else "Profile": nom,
                "Attendu" if L == "fr" else "Expected": attendu,
                "Score": score_str,
            })
        df_diag = pd.DataFrame(rows)
        st.dataframe(df_diag, use_container_width=True, hide_index=True)

        interp_note = (
            f"**Comment lire ce tableau :** Plus la probabilité est élevée, plus le risque d'interruption estimé "
            f"est élevé. Le seuil actif ({SEUIL:.0%}, réglable dans la barre latérale) "
            "détermine à partir de quelle probabilité un patient est classé « à risque »."
            if L == "fr" else
            f"**How to read this table:** The higher the probability, the higher the estimated interruption risk. "
            f"The active threshold ({SEUIL:.0%}, adjustable in the sidebar) determines "
            "from which probability a patient is classified as \"at risk\"."
        )
        st.info(interp_note)

    st.divider()

    seuil_expl = (
        "**Rappel sur le seuil de décision :**\n\n"
        "- **Seuil bas (0,20–0,40)** → Plus de patients signalés (moins de cas à risque manqués), mais plus de fausses alertes\n"
        f"- **Seuil {meta['seuil_defaut']:.2f}** → Seuil recommandé par défaut\n"
        "- **Seuil haut (0,60–0,80)** → Moins de fausses alertes, mais plus de cas à risque potentiellement manqués\n\n"
        "Ajustez le curseur dans la barre latérale selon la priorité clinique de votre structure."
        if L == "fr" else
        "**Decision threshold reminder:**\n\n"
        "- **Low threshold (0.20–0.40)** → More patients flagged (fewer at-risk cases missed), but more false alarms\n"
        f"- **Threshold {meta['seuil_defaut']:.2f}** → Recommended default threshold\n"
        "- **High threshold (0.60–0.80)** → Fewer false alarms, but more at-risk cases potentially missed\n\n"
        "Adjust the slider in the sidebar according to your facility's clinical priority."
    )
    st.info(seuil_expl)


# ─────────────────────────────────────────────────────────────────────────────
# PIED DE PAGE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    f'<div class="footer">{t("footer")}</div>',
    unsafe_allow_html=True,
)
